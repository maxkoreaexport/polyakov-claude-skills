#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl"]
# ///
"""Missed demand analysis for Yandex Direct campaigns.

Subcommands:
  parse-xlsx   Parse Yandex Direct XLSX export into groups/phrases/minus
  build-query  Build OR-query from slot structure
  query-total  Get totalCount from Wordstat API
"""
import argparse
import json
import sys
import urllib.request
import urllib.error
import re
from collections import OrderedDict


# --- Stop words that need + prefix in Wordstat queries ---
STOP_WORDS = frozenset([
    "на", "в", "к", "за", "с", "по", "из", "от", "до", "для",
    "без", "при", "под", "над", "между", "через", "об", "перед",
])

# Characters forbidden in slot variants (OR-syntax operators and leading modifiers)
# Hyphen inside words is OK (б/у, санкт-петербург)
FORBIDDEN_CHARS_RE = re.compile(r'[|()"]')
LEADING_OPERATOR_RE = re.compile(r'^[+\-!]')
# Punctuation to strip from slot variants (dots, commas, semicolons, etc.)
# Keeps hyphens (б/у, санкт-петербург), slashes, and ! (Wordstat exact form operator)
PUNCTUATION_RE = re.compile(r'[.,;:?…·•]')


def cmd_parse_xlsx(args):
    """Parse Yandex Direct XLSX export."""
    import openpyxl

    wb = openpyxl.load_workbook(args.file, data_only=True)
    if "Тексты" not in wb.sheetnames:
        sys.exit(
            f"Error: лист 'Тексты' не найден. Доступные: {wb.sheetnames}"
        )
    ws = wb["Тексты"]

    # Find campaign minus-phrases (scan rows 1-20, cols A-F)
    campaign_minus = ""
    found_minus = False
    for row in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=6, values_only=False):
        if found_minus:
            break
        for cell in row:
            if cell.value and "Минус-фразы на кампанию" in str(cell.value):
                next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                if next_cell.value:
                    campaign_minus = str(next_cell.value).strip()
                found_minus = True
                break

    # Find first data row: col A == "-" and col G non-empty
    data_start = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1, values_only=False):
        cell = row[0]
        if cell.value and str(cell.value).strip() == "-":
            # Check col G
            g_cell = ws.cell(row=cell.row, column=7)
            if g_cell.value and str(g_cell.value).strip():
                data_start = cell.row
                break

    if data_start is None:
        sys.exit("Error: не найдены строки данных (col A='-' с непустой col G)")

    # Collect groups
    groups = OrderedDict()  # group_id -> {name, group_minus_set, phrases_list}
    for row_idx in range(data_start, ws.max_row + 1):
        col_a = ws.cell(row=row_idx, column=1).value
        if col_a is None or str(col_a).strip() != "-":
            continue

        col_g = ws.cell(row=row_idx, column=7).value
        if col_g is None or not str(col_g).strip():
            continue

        col_c = ws.cell(row=row_idx, column=3).value
        if col_c is None or not str(col_c).strip():
            continue  # skip rows with empty group_id

        group_id = str(col_c).strip()
        group_name = str(ws.cell(row=row_idx, column=4).value or "").strip()
        phrase = str(col_g).strip()

        # Col AG (33) - group minus
        col_ag = ws.cell(row=row_idx, column=33).value
        group_minus_val = str(col_ag).strip() if col_ag else ""

        if group_id not in groups:
            groups[group_id] = {
                "name": group_name,
                "group_minus_set": set(),
                "phrases_ordered": OrderedDict(),
            }

        g = groups[group_id]
        # Deduplicate phrases preserving order
        g["phrases_ordered"][phrase] = None
        if group_minus_val:
            g["group_minus_set"].add(group_minus_val)

    wb.close()

    # Build output
    result_groups = []
    for gid, g in groups.items():
        result_groups.append({
            "id": gid,
            "name": g["name"],
            "group_minus": " ".join(sorted(g["group_minus_set"])),
            "phrases": list(g["phrases_ordered"].keys()),
        })

    # Filter by --group if specified
    if args.group:
        result_groups = [g for g in result_groups if g["id"] == args.group]

    result = {
        "campaign_minus": campaign_minus,
        "campaign_minus_length": len(campaign_minus),
        "groups": result_groups,
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))


def sanitize_variant(variant, slot_name):
    """Sanitize a slot variant. Returns (cleaned, warnings)."""
    warnings = []
    cleaned = variant

    # Strip punctuation (муз. центр -> муз центр)
    if PUNCTUATION_RE.search(cleaned):
        cleaned = PUNCTUATION_RE.sub("", cleaned)
        cleaned = re.sub(r'\s+', ' ', cleaned).strip()

    # Remove forbidden OR-syntax characters (replace with space to avoid merging tokens)
    if FORBIDDEN_CHARS_RE.search(cleaned):
        warnings.append(
            f"slot '{slot_name}': вариант '{variant}' содержит OR-синтаксис, деградирован в многословный"
        )
        cleaned = FORBIDDEN_CHARS_RE.sub(" ", cleaned)
        # Collapse multiple spaces
        cleaned = re.sub(r'\s+', ' ', cleaned).strip()

    # Remove leading operators (+ - !) from each token
    tokens = cleaned.split()
    new_tokens = []
    for t in tokens:
        if LEADING_OPERATOR_RE.match(t):
            stripped = LEADING_OPERATOR_RE.sub("", t)
            if stripped:
                warnings.append(
                    f"slot '{slot_name}': убран ведущий оператор из '{t}'"
                )
                new_tokens.append(stripped)
        else:
            new_tokens.append(t)
    cleaned = " ".join(new_tokens)

    return cleaned.strip(), warnings


def add_stop_word_plus(variant):
    """Add + prefix to stop words in a variant string."""
    tokens = variant.split()
    result = []
    for t in tokens:
        # Don't touch tokens already starting with operator
        if t[0] in ("+", "-", "!"):
            result.append(t)
        elif t.lower() in STOP_WORDS:
            result.append(f"+{t}")
        else:
            result.append(t)
    return " ".join(result)


def cmd_build_query(args):
    """Build OR-query from slots JSON."""
    try:
        slots = json.loads(args.slots_json)
    except json.JSONDecodeError as e:
        sys.exit(f"Error: невалидный JSON слотов: {e}")

    max_len = args.max_query_length
    warnings = []
    trimmed = []

    # Slot order: actions, objects, modifiers, additional
    slot_order = ["actions", "objects", "modifiers", "additional"]

    # Sanitize all variants
    sanitized_slots = {}
    for slot_name in slot_order:
        raw_variants = slots.get(slot_name, [])
        if not isinstance(raw_variants, list):
            raw_variants = []
        clean_variants = []
        for v in raw_variants:
            v_str = str(v).strip()
            if not v_str:
                continue
            cleaned, san_warnings = sanitize_variant(v_str, slot_name)
            warnings.extend(san_warnings)
            if cleaned:
                # Apply stop-word plus
                cleaned = add_stop_word_plus(cleaned)
                clean_variants.append(cleaned)
        sanitized_slots[slot_name] = clean_variants

    def build_or_string(slot_variants):
        """Build OR part: (a|b|c) or a."""
        if len(slot_variants) == 0:
            return ""
        if len(slot_variants) == 1:
            return slot_variants[0]
        return "(" + "|".join(slot_variants) + ")"

    def assemble_query(s_slots):
        """Assemble full query from slots dict."""
        parts = []
        for name in slot_order:
            or_str = build_or_string(s_slots[name])
            if or_str:
                parts.append(or_str)
        return " ".join(parts)

    def estimated_phrases(s_slots):
        """Calculate estimated phrase count (product of slot sizes)."""
        product = 1
        for name in slot_order:
            n = len(s_slots[name])
            if n > 0:
                product *= n
        return product

    # Trim priority: additional -> modifiers -> actions (objects untouched)
    trim_order = ["additional", "modifiers", "actions"]

    # Check estimated phrases limit
    while estimated_phrases(sanitized_slots) > 200:
        trimmed_any = False
        for trim_slot in trim_order:
            if len(sanitized_slots[trim_slot]) > 1:
                removed = sanitized_slots[trim_slot].pop()
                trimmed.append({"slot": trim_slot, "removed": removed, "reason": "estimated_phrases > 200"})
                trimmed_any = True
                break
        if not trimmed_any:
            break

    # Check query length limit
    query = assemble_query(sanitized_slots)
    while len(query) > max_len:
        trimmed_any = False
        for trim_slot in trim_order:
            if len(sanitized_slots[trim_slot]) > 1:
                removed = sanitized_slots[trim_slot].pop()
                trimmed.append({"slot": trim_slot, "removed": removed, "reason": f"query_length > {max_len}"})
                trimmed_any = True
                query = assemble_query(sanitized_slots)
                break
        if not trimmed_any:
            warnings.append(
                f"query_length {len(query)} всё ещё > {max_len} после максимальной обрезки"
            )
            break

    result = {
        "query": query,
        "estimated_phrases": estimated_phrases(sanitized_slots),
        "query_length": len(query),
        "trimmed": trimmed,
        "warnings": warnings,
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))


def cmd_query_total(args):
    """Get totalCount from Wordstat API."""
    body = {"phrase": args.phrase}
    if args.regions:
        try:
            body["regions"] = [int(r.strip()) for r in args.regions.split(",")]
        except ValueError:
            sys.exit(f"Error: невалидные regions: {args.regions}")

    data = json.dumps(body).encode("utf-8")
    url = "https://api.wordstat.yandex.net/v1/topRequests"

    req = urllib.request.Request(
        url,
        data=data,
        headers={
            "Authorization": f"Bearer {args.token}",
            "Content-Type": "application/json; charset=utf-8",
        },
        method="POST",
    )

    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            resp_body = resp.read().decode("utf-8")
    except urllib.error.HTTPError as e:
        raw = ""
        try:
            raw = e.read().decode("utf-8")[:500]
        except Exception:
            pass
        result = {"error": f"HTTP {e.code}: {e.reason}", "query": args.phrase, "raw": raw}
        print(json.dumps(result, ensure_ascii=False))
        sys.exit(1)
    except urllib.error.URLError as e:
        result = {"error": f"URL error: {e.reason}", "query": args.phrase}
        print(json.dumps(result, ensure_ascii=False))
        sys.exit(1)

    try:
        obj = json.loads(resp_body)
    except json.JSONDecodeError:
        result = {"error": "Invalid JSON response", "query": args.phrase, "raw": resp_body[:500]}
        print(json.dumps(result, ensure_ascii=False))
        sys.exit(1)

    # Check for errors (multiple possible formats)
    if "error" in obj:
        err_val = obj["error"]
        result = {"error": str(err_val), "query": args.phrase, "raw": resp_body[:500]}
        print(json.dumps(result, ensure_ascii=False))
        sys.exit(1)

    if "error_code" in obj:
        err_msg = f"{obj.get('error_str', '')}: {obj.get('error_detail', '')}"
        result = {
            "error": err_msg.strip(": "),
            "error_code": obj["error_code"],
            "query": args.phrase,
            "raw": resp_body[:500],
        }
        print(json.dumps(result, ensure_ascii=False))
        sys.exit(1)

    # Extract totalCount with fallbacks
    total_count = obj.get("totalCount")
    if total_count is None:
        total_count = (obj.get("result") or {}).get("totalCount")
    if total_count is None:
        result = {
            "error": "totalCount not found in response",
            "query": args.phrase,
            "raw": resp_body[:500],
        }
        print(json.dumps(result, ensure_ascii=False))
        sys.exit(1)

    result = {"total_count": int(total_count), "query": args.phrase}
    print(json.dumps(result, ensure_ascii=False))


def main():
    parser = argparse.ArgumentParser(description="Missed demand analysis tools")
    sub = parser.add_subparsers(dest="command", required=True)

    # parse-xlsx
    p_parse = sub.add_parser("parse-xlsx", help="Parse Yandex Direct XLSX export")
    p_parse.add_argument("file", help="Path to XLSX file")
    p_parse.add_argument("--group", default=None, help="Filter by group ID")

    # build-query
    p_build = sub.add_parser("build-query", help="Build OR-query from slots")
    p_build.add_argument("slots_json", help="JSON string with slots")
    p_build.add_argument("--max-query-length", type=int, default=4096, help="Max query length (default: 4096)")

    # query-total
    p_query = sub.add_parser("query-total", help="Get totalCount from Wordstat API")
    p_query.add_argument("--token", required=True, help="Yandex Wordstat API token")
    p_query.add_argument("--phrase", required=True, help="Search phrase with operators")
    p_query.add_argument("--regions", default=None, help="Region IDs comma-separated")

    args = parser.parse_args()
    if args.command == "parse-xlsx":
        cmd_parse_xlsx(args)
    elif args.command == "build-query":
        cmd_build_query(args)
    elif args.command == "query-total":
        cmd_query_total(args)


if __name__ == "__main__":
    main()
